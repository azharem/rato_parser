from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.views import generic
from openpyxl import load_workbook
import json
from datetime import datetime
from .models import Control, ArsBaseline
from dateutil.parser import *

CONTROL_NUMBER = 0
ARS_BASELINE = 1
CONTROL_TYPE = 2
CONTROL_NAME = 3
CONTROL_SET_VERSION_NUMBER = 4
AUTHORIZATION_PACKAGE_NAME = 5
ALLOCATION_STATUS = 6
OVERALL_UPDATE_DATE = 7
OVERALL_CONTROL_STATUS = 8
ASSESSMENT_STATUS = 9
TRACKING_ID = 10

def process(request):
    print("request received")
    wb = load_workbook('/Users/azhar/Documents/CMS/data/control_tracking_id_sample.xlsx', read_only=True)
    sheet = wb.active
    maxColumns = sheet.max_column
    maxRows = sheet.max_row
    controlDict = {}
    row_number = 2
    controls = []
    ars_baselines = []
    for row in sheet.iter_rows(min_row=2, max_row=maxRows, min_col=1, max_col=maxColumns, values_only=True):
        control = Control()
        control.control_number = row[CONTROL_NUMBER]
        control.control_type = row[CONTROL_TYPE]
        control.control_name = row[CONTROL_NAME]
        control.control_set_version_number = row[CONTROL_SET_VERSION_NUMBER]
        control.authorization_package_name = row[AUTHORIZATION_PACKAGE_NAME]
        control.allocation_status = row[ALLOCATION_STATUS]
        datestr = row[OVERALL_UPDATE_DATE]
        if not datestr:
            control.overall_update_date = None
        else:
            date = datetime.strptime(datestr, '%m/%d/%Y').date()
            control.overall_update_date = date
        control.overall_control_status = row[OVERALL_CONTROL_STATUS]
        control.assessment_status = row[ASSESSMENT_STATUS]
        control.tracking_id = row[TRACKING_ID]
        control.save()
        baselines = row[ARS_BASELINE].split()
        for baseline in baselines:
            newBaseline = ArsBaseline()
            newBaseline.ars_baseline = baseline
            newBaseline.control = control
            newBaseline.save()
    return render(request, 'cmsparser/parse.html',{})

def view(request):
    return render(request,
                  'cmsparser/index.html',
                  {'message' : 'receieved'})

def controls(request):
    controls = Control.objects.all()
    return render(request,
                  'cmsparser/controls.html',
                  {'controls' : controls})
